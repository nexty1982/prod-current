#!/usr/bin/env python3
"""
Compare September 2025 backup with current front-end/src
Generate Excel report with file comparison
"""

import os
import sys
from pathlib import Path
from collections import defaultdict
import json

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)

def get_file_stats(filepath):
    """Get file statistics: size, line count"""
    try:
        size = os.path.getsize(filepath)
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            lines = len(f.readlines())
        return size, lines
    except Exception as e:
        return 0, 0

def get_relative_path(full_path, base_path):
    """Get relative path from base"""
    try:
        return str(Path(full_path).relative_to(base_path))
    except:
        return str(full_path)

def scan_directory(directory):
    """Scan directory and return dict of relative_path -> full_path"""
    files = {}
    base_path = Path(directory).resolve()
    
    for root, dirs, filenames in os.walk(directory):
        # Skip node_modules, .git, dist, build
        dirs[:] = [d for d in dirs if d not in ['node_modules', '.git', 'dist', 'build', '.next']]
        
        for filename in filenames:
            # Skip certain file types
            if filename.startswith('.') and filename not in ['.env', '.gitignore']:
                continue
            
            full_path = os.path.join(root, filename)
            rel_path = get_relative_path(full_path, base_path)
            files[rel_path] = full_path
    
    return files

def compare_directories(backup_dir, current_dir, output_file):
    """Compare two directories and generate Excel report"""
    
    print(f"Scanning backup directory: {backup_dir}")
    backup_files = scan_directory(backup_dir)
    print(f"Found {len(backup_files)} files in backup")
    
    print(f"Scanning current directory: {current_dir}")
    current_files = scan_directory(current_dir)
    print(f"Found {len(current_files)} files in current")
    
    # Categorize files
    only_in_backup = set(backup_files.keys()) - set(current_files.keys())
    only_in_current = set(current_files.keys()) - set(backup_files.keys())
    in_both = set(backup_files.keys()) & set(current_files.keys())
    
    print(f"\nComparison Summary:")
    print(f"  Only in backup: {len(only_in_backup)}")
    print(f"  Only in current: {len(only_in_current)}")
    print(f"  In both: {len(in_both)}")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "File Comparison"
    
    # Headers
    headers = [
        "File Path",
        "Status",
        "Backup Size (bytes)",
        "Backup Lines",
        "Current Size (bytes)",
        "Current Lines",
        "Size Diff (bytes)",
        "Lines Diff",
        "Size Change %",
        "Backup Path",
        "Current Path"
    ]
    
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Fill data
    row_num = 2
    
    # Files only in backup
    for rel_path in sorted(only_in_backup):
        backup_path = backup_files[rel_path]
        size, lines = get_file_stats(backup_path)
        ws.append([
            rel_path,
            "Only in Backup",
            size,
            lines,
            0,
            0,
            -size,
            -lines,
            "-100%",
            backup_path,
            ""
        ])
        row_num += 1
    
    # Files only in current
    for rel_path in sorted(only_in_current):
        current_path = current_files[rel_path]
        size, lines = get_file_stats(current_path)
        ws.append([
            rel_path,
            "Only in Current",
            0,
            0,
            size,
            lines,
            size,
            lines,
            "100%",
            "",
            current_path
        ])
        row_num += 1
    
    # Files in both - compare
    for rel_path in sorted(in_both):
        backup_path = backup_files[rel_path]
        current_path = current_files[rel_path]
        
        backup_size, backup_lines = get_file_stats(backup_path)
        current_size, current_lines = get_file_stats(current_path)
        
        size_diff = current_size - backup_size
        lines_diff = current_lines - backup_lines
        
        if backup_size > 0:
            size_change_pct = f"{(size_diff / backup_size * 100):.1f}%"
        else:
            size_change_pct = "N/A" if current_size == 0 else "100%"
        
        status = "Modified" if size_diff != 0 or lines_diff != 0 else "Unchanged"
        
        ws.append([
            rel_path,
            status,
            backup_size,
            backup_lines,
            current_size,
            current_lines,
            size_diff,
            lines_diff,
            size_change_pct,
            backup_path,
            current_path
        ])
        row_num += 1
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        ws.column_dimensions[column].width = adjusted_width
    
    # Add conditional formatting for status column
    status_col = 2
    for row in range(2, row_num + 1):
        status_cell = ws.cell(row=row, column=status_col)
        if status_cell.value == "Only in Backup":
            status_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        elif status_cell.value == "Only in Current":
            status_cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        elif status_cell.value == "Modified":
            status_cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    
    # Create summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.append(["Comparison Summary"])
    ws_summary.append([])
    ws_summary.append(["Metric", "Count"])
    ws_summary.append(["Total files in backup", len(backup_files)])
    ws_summary.append(["Total files in current", len(current_files)])
    ws_summary.append(["Files only in backup", len(only_in_backup)])
    ws_summary.append(["Files only in current", len(only_in_current)])
    ws_summary.append(["Files in both", len(in_both)])
    
    # Count modified files
    modified_count = sum(1 for rel_path in in_both 
                         if get_file_stats(backup_files[rel_path]) != get_file_stats(current_files[rel_path]))
    ws_summary.append(["Modified files", modified_count])
    ws_summary.append(["Unchanged files", len(in_both) - modified_count])
    
    # Style summary sheet
    ws_summary['A1'].font = Font(bold=True, size=14)
    for cell in ws_summary['A3:B3']:
        cell[0].font = Font(bold=True)
        cell[0].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell[0].font = Font(bold=True, color="FFFFFF")
    
    # Auto-adjust summary column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 15
    
    # Save workbook
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    print(f"\nExcel report saved to: {output_file}")

if __name__ == "__main__":
    # Default paths
    workspace_root = Path(__file__).parent.parent
    backup_dir = workspace_root / "09-25" / "src-9-30-25-working"
    current_dir = workspace_root / "front-end" / "src"
    output_file = workspace_root / "prod" / "docs" / "investigation.xlsx"
    
    # Allow override via command line
    if len(sys.argv) > 1:
        backup_dir = Path(sys.argv[1])
    if len(sys.argv) > 2:
        current_dir = Path(sys.argv[2])
    if len(sys.argv) > 3:
        output_file = Path(sys.argv[3])
    
    if not backup_dir.exists():
        print(f"ERROR: Backup directory not found: {backup_dir}")
        sys.exit(1)
    
    if not current_dir.exists():
        print(f"ERROR: Current directory not found: {current_dir}")
        sys.exit(1)
    
    print(f"Comparing:")
    print(f"  Backup: {backup_dir}")
    print(f"  Current: {current_dir}")
    print(f"  Output: {output_file}")
    print()
    
    compare_directories(str(backup_dir), str(current_dir), str(output_file))
