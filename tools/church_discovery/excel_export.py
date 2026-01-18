"""
Excel export module for church discovery data.

Exports canonical church data to Excel workbook with:
- One sheet per jurisdiction
- Summary sheet with totals
- Formatted headers and columns
"""

from __future__ import annotations
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import CanonicalChurch

logger = logging.getLogger(__name__)

# Column definitions for church data
COLUMNS = [
    ('ID', 'id', 16),
    ('Name', 'name', 40),
    ('Street', 'street', 35),
    ('City', 'city', 20),
    ('State', 'state', 8),
    ('ZIP', 'zip', 12),
    ('Phone', 'phone', 18),
    ('Website', 'website', 45),
    ('Latitude', 'lat', 12),
    ('Longitude', 'lng', 12),
    ('Discovered', 'discovered_at', 20),
]

# Style definitions
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

CELL_ALIGNMENT = Alignment(vertical='top', wrap_text=False)
LINK_FONT = Font(color='0563C1', underline='single')

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def church_to_row(church: CanonicalChurch) -> List[Any]:
    """Convert a church record to a row of values."""
    return [
        church.id,
        church.name,
        church.address.street if church.address else None,
        church.address.city if church.address else None,
        church.address.state if church.address else None,
        church.address.zip if church.address else None,
        church.contact.phone if church.contact else None,
        church.contact.website if church.contact else None,
        church.geo.lat if church.geo else None,
        church.geo.lng if church.geo else None,
        church.timestamps.discovered_at if church.timestamps else None,
    ]


def apply_header_style(ws, row: int = 1) -> None:
    """Apply header styling to the first row."""
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        
        # Set column width
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def add_churches_to_sheet(ws, churches: List[CanonicalChurch], start_row: int = 2) -> int:
    """
    Add church data to a worksheet.
    
    Args:
        ws: Worksheet to add data to
        churches: List of churches to add
        start_row: Row to start adding data (after headers)
        
    Returns:
        Number of rows added
    """
    for row_idx, church in enumerate(churches, start=start_row):
        row_data = church_to_row(church)
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER
            
            # Make website a hyperlink
            if col_idx == 8 and value:  # Website column
                cell.font = LINK_FONT
                try:
                    cell.hyperlink = value
                except Exception:
                    pass  # Skip if hyperlink fails
    
    return len(churches)


def sanitize_sheet_name(name: str) -> str:
    """
    Sanitize a string to be a valid Excel sheet name.
    
    Excel sheet names:
    - Max 31 characters
    - Cannot contain: [ ] : * ? / \
    - Cannot be empty
    """
    if not name:
        return "Unknown"
    
    # Remove invalid characters
    invalid_chars = ['[', ']', ':', '*', '?', '/', '\\']
    sanitized = name
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '')
    
    # Truncate to 31 characters
    if len(sanitized) > 31:
        sanitized = sanitized[:28] + "..."
    
    return sanitized or "Unknown"


def get_jurisdiction_display_name(code: str, full_name: str) -> str:
    """Get a display name for a jurisdiction sheet."""
    # Mapping of codes to short names
    short_names = {
        'goa': 'Greek Orthodox',
        'oca': 'OCA',
        'aocana': 'Antiochian',
        'roc': 'ROCOR',
        'rocor': 'ROCOR',
        'serb': 'Serbian',
        'bulg': 'Bulgarian',
        'rom': 'Romanian',
        'ukr': 'Ukrainian',
        'alb': 'Albanian',
        'carpath': 'Carpatho-Russian',
        'goaa': 'GOAA',
    }
    
    if code and code.lower() in short_names:
        return short_names[code.lower()]
    
    if full_name:
        # Try to create a short version
        if 'Greek Orthodox' in full_name:
            return 'Greek Orthodox'
        if 'Orthodox Church in America' in full_name:
            return 'OCA'
        if 'Antiochian' in full_name:
            return 'Antiochian'
        if 'Russian' in full_name and 'Outside' in full_name:
            return 'ROCOR'
        if 'Ukrainian' in full_name:
            return 'Ukrainian'
        if 'Romanian' in full_name:
            return 'Romanian'
        if 'Serbian' in full_name:
            return 'Serbian'
        if 'Bulgarian' in full_name:
            return 'Bulgarian'
        if 'Albanian' in full_name:
            return 'Albanian'
        if 'Carpatho' in full_name:
            return 'Carpatho-Russian'
        
        # Fallback: use first word(s)
        return sanitize_sheet_name(full_name.split(',')[0][:25])
    
    return code.upper() if code else "Unknown"


def export_to_excel(
    churches: List[CanonicalChurch],
    output_path: str,
    include_summary: bool = True
) -> str:
    """
    Export churches to Excel workbook with one sheet per jurisdiction.
    
    Args:
        churches: List of canonical church records
        output_path: Path to save the Excel file
        include_summary: Whether to include a summary sheet
        
    Returns:
        Path to the created Excel file
    """
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Group churches by jurisdiction
    by_jurisdiction: Dict[str, List[CanonicalChurch]] = {}
    
    for church in churches:
        jur_code = church.jurisdiction_code or 'unknown'
        jur_name = church.jurisdiction or 'Unknown Jurisdiction'
        
        # Create a key that combines code and name for grouping
        key = (jur_code, jur_name)
        
        if key not in by_jurisdiction:
            by_jurisdiction[key] = []
        by_jurisdiction[key].append(church)
    
    # Sort jurisdictions by count (descending)
    sorted_jurisdictions = sorted(
        by_jurisdiction.items(),
        key=lambda x: len(x[1]),
        reverse=True
    )
    
    # Track sheet names to avoid duplicates
    used_sheet_names = set()
    jurisdiction_stats = []
    
    # Create a sheet for each jurisdiction
    for (jur_code, jur_name), jur_churches in sorted_jurisdictions:
        # Get display name for sheet
        sheet_name = get_jurisdiction_display_name(jur_code, jur_name)
        sheet_name = sanitize_sheet_name(sheet_name)
        
        # Handle duplicate sheet names
        original_name = sheet_name
        counter = 1
        while sheet_name in used_sheet_names:
            suffix = f" ({counter})"
            sheet_name = original_name[:31-len(suffix)] + suffix
            counter += 1
        used_sheet_names.add(sheet_name)
        
        # Create worksheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Add headers
        apply_header_style(ws)
        
        # Add data
        count = add_churches_to_sheet(ws, jur_churches)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Track stats
        jurisdiction_stats.append({
            'sheet_name': sheet_name,
            'jurisdiction': jur_name,
            'code': jur_code,
            'count': count
        })
        
        logger.info(f"Added sheet '{sheet_name}' with {count} churches")
    
    # Add summary sheet at the beginning
    if include_summary:
        summary_ws = wb.create_sheet(title="Summary", index=0)
        
        # Title
        summary_ws['A1'] = "Orthodox Church Discovery - Summary"
        summary_ws['A1'].font = Font(bold=True, size=14)
        summary_ws.merge_cells('A1:D1')
        
        # Generation info
        summary_ws['A3'] = "Generated:"
        summary_ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        summary_ws['A4'] = "Total Churches:"
        summary_ws['B4'] = len(churches)
        summary_ws['B4'].font = Font(bold=True)
        
        # Jurisdiction breakdown
        summary_ws['A6'] = "Breakdown by Jurisdiction"
        summary_ws['A6'].font = Font(bold=True, size=12)
        
        # Headers
        summary_ws['A7'] = "Jurisdiction"
        summary_ws['B7'] = "Code"
        summary_ws['C7'] = "Count"
        summary_ws['D7'] = "Sheet"
        
        for col in ['A', 'B', 'C', 'D']:
            cell = summary_ws[f'{col}7']
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
        
        # Data rows
        for idx, stat in enumerate(jurisdiction_stats, start=8):
            summary_ws[f'A{idx}'] = stat['jurisdiction']
            summary_ws[f'B{idx}'] = stat['code']
            summary_ws[f'C{idx}'] = stat['count']
            summary_ws[f'D{idx}'] = stat['sheet_name']
        
        # Total row
        total_row = 8 + len(jurisdiction_stats)
        summary_ws[f'A{total_row}'] = "TOTAL"
        summary_ws[f'A{total_row}'].font = Font(bold=True)
        summary_ws[f'C{total_row}'] = len(churches)
        summary_ws[f'C{total_row}'].font = Font(bold=True)
        
        # Set column widths
        summary_ws.column_dimensions['A'].width = 45
        summary_ws.column_dimensions['B'].width = 10
        summary_ws.column_dimensions['C'].width = 10
        summary_ws.column_dimensions['D'].width = 20
    
    # Save workbook
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb.save(output_path)
    logger.info(f"Excel workbook saved to {output_path}")
    
    return str(output_path)


def export_from_canonical_json(
    json_path: str,
    output_path: str
) -> str:
    """
    Export from a canonical JSON file to Excel.
    
    Args:
        json_path: Path to canonical JSON file
        output_path: Path to save Excel file
        
    Returns:
        Path to created Excel file
    """
    # Load canonical data
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    churches = []
    for church_data in data.get('churches', []):
        church = CanonicalChurch.from_dict(church_data)
        churches.append(church)
    
    logger.info(f"Loaded {len(churches)} churches from {json_path}")
    
    return export_to_excel(churches, output_path)
