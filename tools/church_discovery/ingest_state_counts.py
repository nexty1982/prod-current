#!/usr/bin/env python3
"""
Ingest church-list-US.xlsx and populate us_church_counts table.

Reads all sheets, extracts 2-letter state codes, aggregates counts,
and upserts into orthodoxmetrics_db.us_church_counts.

Idempotent: safe to re-run — uses INSERT ... ON DUPLICATE KEY UPDATE.

Usage:
    python3 tools/church_discovery/ingest_state_counts.py
"""

import os
import sys
from collections import Counter

import openpyxl
import mysql.connector

# --- State normalization mapping (full name → 2-letter code) ---
STATE_NAME_TO_CODE = {
    'alabama': 'AL', 'alaska': 'AK', 'arizona': 'AZ', 'arkansas': 'AR',
    'california': 'CA', 'colorado': 'CO', 'connecticut': 'CT', 'delaware': 'DE',
    'florida': 'FL', 'georgia': 'GA', 'hawaii': 'HI', 'idaho': 'ID',
    'illinois': 'IL', 'indiana': 'IN', 'iowa': 'IA', 'kansas': 'KS',
    'kentucky': 'KY', 'louisiana': 'LA', 'maine': 'ME', 'maryland': 'MD',
    'massachusetts': 'MA', 'michigan': 'MI', 'minnesota': 'MN',
    'mississippi': 'MS', 'missouri': 'MO', 'montana': 'MT', 'nebraska': 'NE',
    'nevada': 'NV', 'new hampshire': 'NH', 'new jersey': 'NJ',
    'new mexico': 'NM', 'new york': 'NY', 'north carolina': 'NC',
    'north dakota': 'ND', 'ohio': 'OH', 'oklahoma': 'OK', 'oregon': 'OR',
    'pennsylvania': 'PA', 'rhode island': 'RI', 'south carolina': 'SC',
    'south dakota': 'SD', 'tennessee': 'TN', 'texas': 'TX', 'utah': 'UT',
    'vermont': 'VT', 'virginia': 'VA', 'washington': 'WA',
    'west virginia': 'WV', 'wisconsin': 'WI', 'wyoming': 'WY',
    'district of columbia': 'DC',
}

VALID_CODES = set(STATE_NAME_TO_CODE.values())


def normalize_state(raw):
    """Normalize a state value to a 2-letter USPS code."""
    if not raw:
        return None
    s = str(raw).strip()
    # Already a 2-letter code?
    if len(s) == 2 and s.upper() in VALID_CODES:
        return s.upper()
    # Try full name lookup
    code = STATE_NAME_TO_CODE.get(s.lower())
    if code:
        return code
    return None


def read_excel(path):
    """Read all sheets and return a Counter of state_code → count."""
    wb = openpyxl.load_workbook(path, read_only=True)
    counts = Counter()
    skipped = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find state column index from header row
        header = [str(c).strip().lower() if c else '' for c in rows[0]]
        try:
            state_idx = header.index('state')
        except ValueError:
            print(f"  WARNING: Sheet '{sheet_name}' has no 'State' column, skipping")
            continue

        for row in rows[1:]:
            if state_idx >= len(row):
                skipped += 1
                continue
            code = normalize_state(row[state_idx])
            if code:
                counts[code] += 1
            else:
                skipped += 1

    wb.close()
    print(f"  Total churches with valid state: {sum(counts.values())}")
    print(f"  Skipped (no/invalid state): {skipped}")
    print(f"  States covered: {len(counts)}")
    return counts


def upsert_counts(counts, db_config):
    """Upsert state counts into us_church_counts table."""
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor()

    sql = """
        INSERT INTO us_church_counts (state_code, church_count)
        VALUES (%s, %s)
        ON DUPLICATE KEY UPDATE
            church_count = VALUES(church_count)
    """

    for state_code, count in sorted(counts.items()):
        cursor.execute(sql, (state_code, count))

    conn.commit()
    print(f"  Upserted {len(counts)} state rows into us_church_counts")
    cursor.close()
    conn.close()


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(script_dir, 'church-list-US.xlsx')

    if not os.path.exists(xlsx_path):
        print(f"ERROR: Excel file not found at {xlsx_path}")
        sys.exit(1)

    print(f"Reading {xlsx_path} ...")
    counts = read_excel(xlsx_path)

    if not counts:
        print("ERROR: No valid state data found")
        sys.exit(1)

    # Print summary
    print("\nState counts:")
    for code in sorted(counts):
        print(f"  {code}: {counts[code]}")

    # DB config from environment or defaults
    db_config = {
        'host': os.environ.get('DB_HOST', 'localhost'),
        'user': os.environ.get('DB_USER', 'orthodoxapps'),
        'password': os.environ.get('DB_PASSWORD', 'Summerof1982@!'),
        'database': os.environ.get('DB_NAME', 'orthodoxmetrics_db'),
        'port': int(os.environ.get('DB_PORT', '3306')),
    }

    print(f"\nUpserting into {db_config['database']}.us_church_counts ...")
    upsert_counts(counts, db_config)
    print("Done!")


if __name__ == '__main__':
    main()
